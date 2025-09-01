import win32com.client as win32
import openpyxl
import time

if __name__ == '__main__':

    # 加载Excel工作簿
    wb = openpyxl.load_workbook('D:\\工作\\主要事项\\2025Q2\\USB\\邮件发送\\申请理由发送总表-内邮.xlsx')
    ws = wb.active
    step = 0

    # 初始化Outlook应用实例
    outlook = win32.Dispatch('outlook.application')

    # 获取当前登录账户的替代方法
    try:
        # 尝试获取当前账户的邮件地址
        namespace = outlook.GetNamespace("MAPI")
        current_user = namespace.CurrentUser.Address
        print(f"当前登录账户: {current_user}")
    except Exception as e:
        current_user = "未知账户"
        print(f"获取登录账户失败: {e}, 将使用默认发件人")

    # 遍历Excel第2行到第4行
    for row in range(2, 34):
        step += 1

        # 新建邮件对象
        mail = outlook.CreateItem(0)

        # 从Excel获取收件人邮箱（第1列）
        mailOut = ws.cell(row, 1).value
        # 构建附件路径
        attachmentOut = 'D:\\工作\\主要事项\\2025Q2\\USB\\邮件发送\\附件\\外设权限开放清单-' + ws.cell(row, 2).value + '.xlsx'

        mail.To = mailOut
        mail.Subject = '关于员工终端外设权限开放确认事项'
        mail.BodyFormat = 2  # HTML格式
        mail.HTMLBody = '''<p style="line-height:2.0em;"><Font Face=微软雅黑 Size=3>亲爱的同事，您好：<br>
                            &ensp;&ensp;&ensp;&ensp;依据《办公终端安全管理规范》5.1 安全管理原则，员工办公终端外设权限开放应遵循最小化外设使用原则，严格控制终端外设的使用需求。<br>
                            &ensp;&ensp;&ensp;&ensp;请您查看邮件附件的外设权限开放详情，如因工作必要，烦请在<Font Color=red><b>本周五前</b></font>访问云文档说明原因；（https://docs.hikvision.com/#/file/nodcnhuV4N242roMypbNdi15uL5）<br>
                            &ensp;&ensp;&ensp;&ensp;如非工作必要，我们将于<Font Color=red><b>本周五统一关闭外设权限，</b></font>感谢您的理解和支持！<br>
                            &ensp;&ensp;&ensp;&ensp;如有疑问，可随时咨询网络安全部习金来，谢谢！<br>
                            &ensp;&ensp;&ensp;&ensp;--------------------------------------------------------------------------------------------------------------------------<br>
                            <Font Size=2.5>本邮件及其附件含有海康威视公司的保密信息，仅限于发送给上面地址中列出的个人或群组。禁止任何其他人以任何形式使用（包括但不限于全部或部分地泄露、复制、或散发）本邮件中的信息。如果您错收了本邮件，请您立即电话或邮件通知发件人并删除本邮件！This e-mail and its attachments contain confidential information from HIKVISION, which is intended only for the person or entity whose address is listed above. Any use of the information contained herein in any way(including, but not limited to, total or partial disclosure, reproduction, or dissemination) by persons other than the intended recipient(s) is prohibited. If you receive this e-mail in error, please notify the sender byphone or email immediately and delete it!'''

        mail.Attachments.Add(attachmentOut)
        mail.Send()  # 发送邮件

        print(f"{step}. 邮件已发送至: {mailOut}")
        time.sleep(3)

    wb.close()
    print("所有邮件发送完成！")
